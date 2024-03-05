import requests
from requests.packages.urllib3.exceptions import InsecureRequestWarning
import json
from switch_ssh_cli import _open_session, _issue_command, _close_session
from regex_library import _express_select
import os
import openpyxl
import socket

# 移除警告
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

class USER_Info(object):
    def _login_info(self):
        self.NCM = ""
        self.ad_username = ""
        self.ad_password = ""
        self.r_session = requests.session()

    def _path_info(self):
        self.dir_path = os.path.abspath(os.path.dirname(os.path.abspath(__file__)) + os.path.sep + ".") + "/"

class SSH_Device(USER_Info):
    def _collect_data(self, device_name, command_list, device_type):
        self._login_info()
        self._path_info()
        output_dict = {}
        ssh_session = _open_session(self.ad_username, self.ad_password, self.ad_username, self.ad_password,"on",self.NCM,device_name,"enable")
        for command in command_list:
            session_output = _issue_command(ssh_session, device_name, command, "enable")
            output = _express_select(session_output,self.dir_path,command,device_type)
            if output !=[]:
                output_dict[command] = output

        _close_session(ssh_session,"enable")
        return output

class API_query(USER_Info):
    def _post_data(self, headers, url):
        USER_Info._login_info(self)
        payload = {
            "aaaUser": {
                "attributes": {
                    "name": "apic#tacacs\|||{0}".format(self.ad_username),
                    "pwd": "{0}".format(self.ad_password)
                }
            }
        }
        data = json.dumps(payload)
        response_data = self.r_session.post(url=url, data=data, headers=headers, verify=False, allow_redirects=False, timeout=5)
        return self.r_session, response_data.status_code

    def _get_data(self, headers, uri, r_session):
        USER_Info._login_info(self)
        response_data = r_session.get(url=uri, verify=False, allow_redirects=False, headers=headers, timeout=10)
        return r_session, response_data.status_code, response_data


class Excel_operate(USER_Info):
    def _excel_create(self, file_name):
        # 创建 excel 文件
        wb = openpyxl.Workbook()
        wb.save(file_name)
        print("-> {0} has been created".format(file_name))
        return wb

    def _write_summary_info(self, wb, file_name, summary_info_list):
        ws = wb.active
        ws.title = "full IP & Mac record"
        title_style = openpyxl.styles.Font(u'等线', size=11, bold=True)
        fill_style = openpyxl.styles.PatternFill(patternType='solid', fgColor="99cc99")
        font_style = openpyxl.styles.Font(u'等线', size=11)
        # ...（前面的代码）

        # 设置列宽
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 90
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 10
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 15
        ws.column_dimensions["J"].width = 15

        ws.append(["Switch", "Port", "Port Description", "Port Status", "Duplex", "Speed", "Vlan", "Mac", "Cat6k IP","ACI IP"])
        # 设置标题样式
        for column in range(11):
            ws.cell(column=column + 1, row=1).font = title_style

        # 将数据写入表格
        for index, each_entry in enumerate(summary_info_list):
            ws.append(each_entry)
            row = index + 2
            for col in range(10):  # 根据你的数据有10列
                ws.cell(column=col + 1, row=row).font = font_style
        
        wb.save(filename = file_name)
        print("-> {0} saved within IP & Mac info...".format(file_name))
        
    def write_server_info(self, wb, file_name, server_info_list):
        title_style = openpyxl.styles.Font(u'等线', size=11, bold=True)
        fill_style = openpyxl.styles.PatternFill(patternType='solid', fgColor='99cc99')
        font_style = openpyxl.styles.Font(u'等线', size=11)
        
        # 创建 sheet
        wb.create_sheet("Server connections")
        ws = wb["Server connections"]
        
        # 设置列宽
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 90
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 10

        # 设置标题样式
        ws.append(["Switch", "Port", "Remark", "Port Description", "Port Status", "Duplex", "Speed"])
        for column in range(7):
            ws.cell(column=column + 1, row=1).font = title_style

        server_index_list = []
        server_identify_list = []

        for each_entry in server_info_list:
            server_identify = each_entry[2]
            if server_identify not in server_identify_list:
                server_identify_list.append(server_identify)

        for remark in server_identify_list:
            single_server_index = [i+2 for i, x in enumerate(server_info_list) if x[2] == remark]
            for index in single_server_index:
                server_index_list.append(index)
        
        for each_entry in server_index_list:
            ws.append(each_entry)
        
        count = 0
        for single_server_index in server_index_list:
            count += 1
            if (count % 2) == 0:
                for index_number in single_server_index:
                    for column in range(7):
                        ws.cell(column=column + 1, row=index_number).fill = fill_style
                        ws.cell(column=column + 1, row=index_number).font = font_style
            else:
                for index_number in single_server_index:
                    for column in range(7):
                        ws.cell(column=column + 1, row=index_number).font = font_style

        wb.save(filename = file_name)
        print("-> {0} saved within server info...".format(file_name))

# ... (前面的代码省略)

    def _write_owner_info(self, wb, file_name, summary_info_list):
        headers = {'content-type': 'application/json', 'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x'}

        # 创建 sheet
        wb.create_sheet("Service owner")
        ws = wb["Service owner"]

        # 设置列宽
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 20

        title_style = openpyxl.styles.Font(name='等线', size=11, bold=True)
        fill_style = openpyxl.styles.PatternFill(patternType='solid', fgColor='99CC99')
        font_style = openpyxl.styles.Font(name='等线', size=11)

        # 设置标题样式
        ws.append(["IP", "FQDN", "Server", "Service name", "Service level", "GSD Group", "Service owner (list)", "Ser"])
        for column in range(7):
            ws.cell(column=column + 1, row=1).font = title_style

        for index, each_entry in enumerate(summary_info_list):
            ws.append(each_entry)
            row = index + 2
            for column in range(7):
                ws.cell(column=column + 1, row=row).font = font_style
                ws.cell(column=column + 1, row=row).fill = fill_style

        # 保存文件
        wb.save(filename=file_name)
        print("-> {0} saved within owner info...".format(file_name))
    









